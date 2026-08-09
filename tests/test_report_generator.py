"""Tests for Excel report generation."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.models import CleaningReport, FileInfo, ValidationResult
from src.report_generator import generate_report

EXPECTED_SHEETS = [
    "Dashboard",
    "Executive Summary",
    "Monthly Performance",
    "Product Performance",
    "Salesperson Performance",
    "Regional Performance",
    "Customer Performance",
    "Data Quality",
    "Raw Data",
]


def _minimal_validation() -> ValidationResult:
    return ValidationResult(valid=True, rows_checked=3, rows_rejected=0)


def _minimal_cleaning() -> CleaningReport:
    return CleaningReport(
        rows_before=4,
        rows_after_duplicates=3,
        duplicate_rows_removed=1,
        invalid_rows_removed=0,
        missing_customer_names_filled=1,
        missing_discounts_filled=1,
        rows_final=3,
    )


def test_report_creates_all_sheets(config, clean_df, raw_df):
    from src.analytics import compute_all

    from src.data_cleaner import clean_dataframe

    file_infos = [FileInfo(path=Path("sales_2026_01.xlsx"), rows_loaded=len(raw_df), sheet_used="Sales")]
    cleaned, cleaning = clean_dataframe(raw_df)
    analytics = compute_all(cleaned)
    validation = _minimal_validation()

    path = generate_report(
        config, cleaned, file_infos, validation, cleaning, analytics
    )
    assert path.exists()
    assert path.stat().st_size > 0

    wb = load_workbook(path, read_only=False)
    assert wb.sheetnames == EXPECTED_SHEETS


def test_dashboard_kpis(config, clean_df):
    from src.analytics import compute_all

    file_infos = [FileInfo(path=Path("sales_2026_01.xlsx"), rows_loaded=len(clean_df), sheet_used="Sales")]
    analytics = compute_all(clean_df)
    path = generate_report(
        config, clean_df, file_infos, _minimal_validation(), _minimal_cleaning(), analytics
    )
    wb = load_workbook(path, data_only=True)
    ws = wb["Dashboard"]
    # KPI cards: first card (Total Revenue) has label at B4 and value at B5.
    assert ws["B4"].value == "Total Revenue"
    assert ws["B5"].value == analytics.metrics["total_revenue"]


def test_data_quality_sheet_populated(config, clean_df):
    from src.analytics import compute_all

    file_infos = [FileInfo(path=Path("sales_2026_01.xlsx"), rows_loaded=4, sheet_used="Sales")]
    analytics = compute_all(clean_df)
    path = generate_report(
        config, clean_df, file_infos, _minimal_validation(), _minimal_cleaning(), analytics
    )
    wb = load_workbook(path, data_only=True)
    ws = wb["Data Quality"]
    # "Rows Included in Final Report" label + value.
    values = [ws.cell(row=r, column=1).value for r in range(1, 40)]
    assert "Rows Included in Final Report" in values


def test_raw_data_sheet_contains_transactions(config, clean_df):
    from src.analytics import compute_all

    file_infos = [FileInfo(path=Path("sales_2026_01.xlsx"), rows_loaded=len(clean_df), sheet_used="Sales")]
    analytics = compute_all(clean_df)
    path = generate_report(
        config, clean_df, file_infos, _minimal_validation(), _minimal_cleaning(), analytics
    )
    wb = load_workbook(path, data_only=True)
    ws = wb["Raw Data"]
    # Header row 4, data starts row 5. clean_df has 3 rows.
    assert ws["A4"].value == "Order ID"
    assert ws["A5"].value == "NS-2026-00001"
    assert ws["A7"].value == "NS-2026-00003"


def test_dashboard_contains_charts(config, clean_df):
    from src.analytics import compute_all

    file_infos = [FileInfo(path=Path("sales_2026_01.xlsx"), rows_loaded=len(clean_df), sheet_used="Sales")]
    analytics = compute_all(clean_df)
    path = generate_report(
        config, clean_df, file_infos, _minimal_validation(), _minimal_cleaning(), analytics
    )
    wb = load_workbook(path)
    ws = wb["Dashboard"]
    assert len(ws._charts) >= 4
