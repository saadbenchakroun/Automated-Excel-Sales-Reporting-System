"""Professional multi-sheet Excel report generation with openpyxl."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config import Config
from src.models import CleaningReport, FileInfo, ValidationResult, AnalyticsResult

logger = logging.getLogger("sales_reporting.report")

NAVY = "1F4E79"
WHITE = "FFFFFF"
LIGHT_BLUE = "DCE6F1"
LIGHT_GRAY = "F2F2F2"
GOLD = "BF8F00"
DARK_TEXT = "333333"

THIN_SIDE = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color=DARK_TEXT)
SECTION_FONT = Font(name="Calibri", size=13, bold=True, color=NAVY)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BODY_FONT = Font(name="Calibri", size=11, color=DARK_TEXT)
KPI_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
KPI_LABEL_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
KPI_VALUE_FONT = Font(name="Calibri", size=15, bold=True, color=NAVY)
ALT_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)


def _currency_format(config: Config) -> str:
    return f'"{config.currency_symbol}"#,##0.00'


def _set_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_title(ws, company: str, subtitle: str, end_col: int = 12) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=1, value=company)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18


def _write_headers(ws, row: int, headers: list[str], start_col: int = 1) -> int:
    for offset, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + offset, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 1


def _apply_borders(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def _write_table(
    ws,
    start_row: int,
    headers: list[str],
    rows: list[list],
    start_col: int = 1,
    number_formats: dict[int, str] | None = None,
    formats: list[str] | None = None,
) -> int:
    """Write a header + data table. Returns the row after the last data row."""
    first_data = _write_headers(ws, start_row, headers, start_col)
    end_col = start_col + len(headers) - 1
    formats = formats or ([None] * len(headers))
    for i, row in enumerate(rows):
        r = first_data + i
        for j, value in enumerate(row):
            cell = ws.cell(row=r, column=start_col + j, value=value)
            cell.font = BODY_FONT
            if formats[j]:
                cell.number_format = formats[j]
    _apply_borders(ws, start_row, first_data + len(rows) - 1, start_col, end_col)
    return first_data + len(rows)


def _add_autofilter(ws, start_row: int, headers: list[str], data_rows: int) -> None:
    end_row = start_row + data_rows
    end_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{start_row}:{end_col}{end_row}"


def _freeze_header(ws) -> None:
    ws.freeze_panes = "A2"


def _kpi_card(ws, row: int, col: int, label: str, value, fmt: str) -> None:
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = KPI_LABEL_FONT
    label_cell.fill = KPI_LABEL_FILL
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = THIN_BORDER

    value_cell = ws.cell(row=row + 1, column=col, value=value)
    value_cell.font = KPI_VALUE_FONT
    value_cell.number_format = fmt
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.border = THIN_BORDER


# ---------------------------------------------------------------- sheets

def _build_dashboard(
    ws,
    config: Config,
    analytics: AnalyticsResult,
) -> None:
    ws.sheet_view.showGridLines = False
    metrics = analytics.metrics
    period = f"{metrics['period_start']} to {metrics['period_end']}"
    _write_title(ws, config.company_name, f"Sales Dashboard  |  Period: {period}", end_col=16)

    currency = _currency_format(config)
    kpis = [
        ("Total Revenue", metrics["total_revenue"], currency),
        ("Total Orders", metrics["total_orders"], "#,##0"),
        ("Avg Order Value", metrics["avg_order_value"], currency),
        ("Units Sold", metrics["total_units"], "#,##0"),
        ("Customers", metrics["total_customers"], "#,##0"),
        ("Cancellation Rate", metrics["cancellation_rate"] / 100.0, "0.00%"),
    ]
    for i, (label, value, fmt) in enumerate(kpis):
        _kpi_card(ws, 4, 2 + i, label, value, fmt)
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 26

    # Data tables the charts reference (kept visible on the left of the dashboard).
    monthly = analytics.monthly
    cat = analytics.by_category
    region = analytics.by_region
    top10 = analytics.top_products

    month_labels = [str(m) for m in monthly["month_label"]]
    month_revenue = [float(v) for v in monthly["net_revenue"]]
    next_row = _write_table(
        ws,
        8,
        ["Month", "Revenue"],
        list(zip(month_labels, month_revenue)),
        start_col=2,
        formats=[None, currency],
    )
    end_month = next_row - 1

    cat_rows = [[str(r["Category"]), float(r["net_revenue"])] for _, r in cat.iterrows()]
    next_row = _write_table(
        ws,
        16,
        ["Category", "Revenue"],
        cat_rows,
        start_col=2,
        formats=[None, currency],
    )
    end_cat = next_row - 1

    region_rows = [[str(r["Region"]), float(r["net_revenue"])] for _, r in region.iterrows()]
    next_row = _write_table(
        ws,
        22,
        ["Region", "Revenue"],
        region_rows,
        start_col=2,
        formats=[None, currency],
    )
    end_region = next_row - 1

    top_rows = [[str(r["Product Name"]), float(r["net_revenue"])] for _, r in top10.iterrows()]
    next_row = _write_table(
        ws,
        29,
        ["Top 10 Products", "Revenue"],
        top_rows,
        start_col=2,
        formats=[None, currency],
    )
    end_top = next_row - 1

    def _ref(top_left: str, bottom_right: str) -> Reference:
        return Reference(ws, min_col=ws[top_left].column, max_col=ws[bottom_right].column,
                         min_row=ws[top_left].row, max_row=ws[bottom_right].row)

    # Charts arranged in a 2x2 grid. Left column (E) sits right next to the data
    # tables; right column (L) with a spacer (K). Row 2 starts at row 28 so the
    # ~7.5 cm (14-row) top charts never overlap the bottom row.
    line = LineChart()
    line.title = "Monthly Revenue Trend"
    line.style = 2
    line.y_axis.title = "Revenue"
    line.add_data(_ref(f"C{9}", f"C{end_month}"), titles_from_data=False)
    line.set_categories(_ref(f"B{9}", f"B{end_month}"))
    line.width, line.height = 14, 7.5
    ws.add_chart(line, "E8")

    cat_chart = BarChart()
    cat_chart.title = "Revenue by Category"
    cat_chart.style = 10
    cat_chart.add_data(_ref(f"C{17}", f"C{end_cat}"), titles_from_data=False)
    cat_chart.set_categories(_ref(f"B{17}", f"B{end_cat}"))
    cat_chart.legend = None
    cat_chart.width, cat_chart.height = 13, 6
    ws.add_chart(cat_chart, "L8")

    region_chart = BarChart()
    region_chart.title = "Revenue by Region"
    region_chart.style = 10
    region_chart.add_data(_ref(f"C{23}", f"C{end_region}"), titles_from_data=False)
    region_chart.set_categories(_ref(f"B{23}", f"B{end_region}"))
    region_chart.legend = None
    region_chart.width, region_chart.height = 13, 6
    ws.add_chart(region_chart, "E28")

    top_chart = BarChart()
    top_chart.type = "bar"
    top_chart.title = "Top 10 Products"
    top_chart.style = 10
    top_chart.add_data(_ref(f"C{30}", f"C{end_top}"), titles_from_data=False)
    top_chart.set_categories(_ref(f"B{30}", f"B{end_top}"))
    top_chart.legend = None
    top_chart.width, top_chart.height = 14, 7.5
    ws.add_chart(top_chart, "L28")

    _set_widths(ws, {
        "A": 2, "B": 20, "C": 16, "D": 3,
        "E": 13, "F": 13, "G": 13, "H": 13, "I": 13, "J": 13,
        "K": 3, "L": 13, "M": 13, "N": 13, "O": 13, "P": 13, "Q": 13,
    })


def _build_executive_summary(
    ws,
    config: Config,
    analytics: AnalyticsResult,
    file_infos: list[FileInfo],
) -> None:
    ws.sheet_view.showGridLines = False
    metrics = analytics.metrics
    _write_title(ws, "Executive Summary", f"{config.company_name} | Sales Reporting Automation", end_col=6)

    period = f"{metrics['period_start']} to {metrics['period_end']}"
    ok_files = [f for f in file_infos if f.status == "ok"]
    info_rows = [
        ["Company", config.company_name],
        ["Report Period", period],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Input Files Processed", len(ok_files)],
        ["Total Rows Loaded", sum(f.rows_loaded for f in ok_files)],
    ]
    currency = _currency_format(config)
    metric_rows = [
        ["Total Revenue (Net)", metrics["total_revenue"], currency],
        ["Gross Revenue", metrics["gross_revenue"], currency],
        ["Total Discounts Given", metrics["total_discounts"], currency],
        ["Number of Orders", metrics["total_orders"], "#,##0"],
        ["Units Sold", metrics["total_units"], "#,##0"],
        ["Average Order Value", metrics["avg_order_value"], currency],
        ["Average Units per Order", metrics["avg_units_per_order"], "0.00"],
        ["Number of Customers", metrics["total_customers"], "#,##0"],
        ["Number of Products Sold", metrics["total_products"], "#,##0"],
        ["Cancellation Rate", metrics["cancellation_rate"] / 100.0, "0.00%"],
    ]
    highlight_rows = [
        ["Top Category", metrics["top_category"]],
        ["Top Region", metrics["top_region"]],
        ["Top Salesperson", metrics["top_salesperson"]],
        ["Top Product", metrics["top_product"]],
    ]

    section = ws.cell(row=4, column=1, value="Key Metrics")
    section.font = SECTION_FONT
    row = 6
    for label, value, fmt in metric_rows:
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = BODY_FONT
        cell.number_format = fmt
        row += 1
    _apply_borders(ws, 6, row - 1, 1, 2)

    ws.cell(row=row + 1, column=1, value="Report Information").font = SECTION_FONT
    _write_table(ws, row + 2, ["Field", "Detail"], info_rows, start_col=1)

    ws.cell(row=row + 2 + len(info_rows) + 2, column=1, value="Highlights").font = SECTION_FONT
    _write_table(ws, row + 2 + len(info_rows) + 3, ["Metric", "Leader"], highlight_rows, start_col=1)

    _set_widths(ws, {"A": 26, "B": 24, "C": 18})


def _build_monthly(ws, analytics: AnalyticsResult, config: Config) -> None:
    _write_title(ws, "Monthly Performance", f"{config.company_name} | Revenue, orders and growth by month", end_col=7)
    currency = _currency_format(config)
    monthly = analytics.monthly
    headers = ["Month", "Orders", "Units", "Gross Revenue", "Discounts", "Net Revenue", "MoM Growth %"]
    rows = []
    for _, r in monthly.iterrows():
        rows.append([
            str(r["month_label"]),
            int(r["orders"]),
            int(r["units"]),
            float(r["gross_revenue"]),
            float(r["discounts"]),
            float(r["net_revenue"]),
            float(r["mom_growth_pct"]),
        ])
    first_data = _write_headers(ws, 4, headers)
    formats = [None, "#,##0", "#,##0", currency, currency, currency, '0.00"%"']
    for i, row in enumerate(rows):
        r = first_data + i
        for j, value in enumerate(row):
            cell = ws.cell(row=r, column=j + 1, value=value)
            cell.font = BODY_FONT
            if formats[j]:
                cell.number_format = formats[j]
    end_row = first_data + len(rows) - 1
    _apply_borders(ws, 4, end_row, 1, len(headers))
    ws.conditional_formatting.add(
        f"G{first_data}:G{end_row}",
        ColorScaleRule(
            start_type="num", start_value=-100, start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFEB84",
            end_type="num", end_value=100, end_color="63BE7B",
        ),
    )
    _add_autofilter(ws, 4, headers, len(rows))
    _freeze_header(ws)
    _set_widths(ws, {"A": 14, "B": 10, "C": 10, "D": 16, "E": 16, "F": 16, "G": 14})


def _build_product(ws, analytics: AnalyticsResult, config: Config) -> None:
    _write_title(ws, "Product Performance", f"{config.company_name} | Revenue by product", end_col=7)
    currency = _currency_format(config)
    headers = ["Product ID", "Product Name", "Category", "Orders", "Units", "Net Revenue", "% of Revenue"]
    data = analytics.top_products if len(analytics.top_products) else analytics.by_product
    rows = []
    for _, r in data.iterrows():
        rows.append([
            r.get("Product ID", ""),
            str(r["Product Name"]),
            str(r.get("Category", "")),
            int(r["orders"]),
            int(r["units"]),
            float(r["net_revenue"]),
            float(r["revenue_share"]),
        ])
    first_data = _write_headers(ws, 4, headers)
    formats = [None, None, None, "#,##0", "#,##0", currency, "0.00%"]
    for i, row in enumerate(rows):
        r = first_data + i
        for j, value in enumerate(row):
            cell = ws.cell(row=r, column=j + 1, value=value)
            cell.font = BODY_FONT
            if formats[j]:
                cell.number_format = formats[j]
    end_row = first_data + len(rows) - 1
    _apply_borders(ws, 4, end_row, 1, len(headers))
    ws.conditional_formatting.add(
        f"F{first_data}:F{end_row}",
        DataBarRule(start_type="min", end_type="max", color="638EC6", showValue=True),
    )
    _add_autofilter(ws, 4, headers, len(rows))
    _freeze_header(ws)
    _set_widths(ws, {"A": 14, "B": 30, "C": 18, "D": 10, "E": 10, "F": 16, "G": 14})


def _build_salesperson(ws, analytics: AnalyticsResult, config: Config) -> None:
    _write_title(ws, "Salesperson Performance", f"{config.company_name} | Revenue by salesperson", end_col=5)
    currency = _currency_format(config)
    headers = ["Salesperson", "Orders", "Units", "Net Revenue", "% of Revenue"]
    rows = []
    for _, r in analytics.by_salesperson.iterrows():
        rows.append([
            str(r["Salesperson"]),
            int(r["orders"]),
            int(r["units"]),
            float(r["net_revenue"]),
            float(r["revenue_share"]),
        ])
    first_data = _write_headers(ws, 4, headers)
    formats = [None, "#,##0", "#,##0", currency, "0.00%"]
    for i, row in enumerate(rows):
        r = first_data + i
        for j, value in enumerate(row):
            cell = ws.cell(row=r, column=j + 1, value=value)
            cell.font = BODY_FONT
            if formats[j]:
                cell.number_format = formats[j]
    end_row = first_data + len(rows) - 1
    _apply_borders(ws, 4, end_row, 1, len(headers))
    _add_autofilter(ws, 4, headers, len(rows))
    _freeze_header(ws)
    _set_widths(ws, {"A": 22, "B": 10, "C": 10, "D": 16, "E": 14})


def _build_region(ws, analytics: AnalyticsResult, config: Config) -> None:
    _write_title(ws, "Regional Performance", f"{config.company_name} | Revenue by region", end_col=5)
    currency = _currency_format(config)
    headers = ["Region", "Orders", "Units", "Net Revenue", "% of Revenue"]
    rows = []
    for _, r in analytics.by_region.iterrows():
        rows.append([
            str(r["Region"]),
            int(r["orders"]),
            int(r["units"]),
            float(r["net_revenue"]),
            float(r["revenue_share"]),
        ])
    first_data = _write_headers(ws, 4, headers)
    formats = [None, "#,##0", "#,##0", currency, "0.00%"]
    for i, row in enumerate(rows):
        r = first_data + i
        for j, value in enumerate(row):
            cell = ws.cell(row=r, column=j + 1, value=value)
            cell.font = BODY_FONT
            if formats[j]:
                cell.number_format = formats[j]
    end_row = first_data + len(rows) - 1
    _apply_borders(ws, 4, end_row, 1, len(headers))
    _add_autofilter(ws, 4, headers, len(rows))
    _freeze_header(ws)
    _set_widths(ws, {"A": 22, "B": 10, "C": 10, "D": 16, "E": 14})


def _build_customer(ws, analytics: AnalyticsResult, config: Config) -> None:
    _write_title(ws, "Customer Performance", f"{config.company_name} | Top 10 customers by revenue", end_col=6)
    currency = _currency_format(config)
    headers = ["Customer ID", "Customer Name", "Orders", "Units", "Net Revenue", "% of Revenue"]
    rows = []
    for _, r in analytics.top_customers.iterrows():
        rows.append([
            str(r["Customer ID"]),
            str(r["Customer Name"]),
            int(r["orders"]),
            int(r["units"]),
            float(r["net_revenue"]),
            float(r["revenue_share"]),
        ])
    first_data = _write_headers(ws, 4, headers)
    formats = [None, None, "#,##0", "#,##0", currency, "0.00%"]
    for i, row in enumerate(rows):
        r = first_data + i
        for j, value in enumerate(row):
            cell = ws.cell(row=r, column=j + 1, value=value)
            cell.font = BODY_FONT
            if formats[j]:
                cell.number_format = formats[j]
    end_row = first_data + len(rows) - 1
    _apply_borders(ws, 4, end_row, 1, len(headers))
    _add_autofilter(ws, 4, headers, len(rows))
    _freeze_header(ws)
    _set_widths(ws, {"A": 14, "B": 30, "C": 10, "D": 10, "E": 16, "F": 14})


def _build_data_quality(
    ws,
    config: Config,
    file_infos: list[FileInfo],
    validation: ValidationResult,
    cleaning: CleaningReport,
) -> None:
    _write_title(ws, "Data Quality Report", f"{config.company_name} | What the pipeline found and fixed", end_col=5)

    ws.cell(row=4, column=1, value="Input Files").font = SECTION_FONT
    file_headers = ["File", "Rows Loaded", "Worksheet", "Status", "Message"]
    file_rows = [
        [info.path.name, info.rows_loaded, info.sheet_used, info.status, info.message]
        for info in file_infos
    ]
    _write_table(ws, 5, file_headers, file_rows)

    ws.cell(row=6 + len(file_rows), column=1, value="Validation Summary").font = SECTION_FONT
    validation_rows = [
        ["Rows Checked", validation.rows_checked],
        ["Validation Errors", validation.error_count],
        ["Validation Warnings", validation.warning_count],
        ["Rows Rejected (>=1 error)", validation.rows_rejected],
        ["Dataset Valid", "Yes" if validation.valid else "No"],
    ]
    _write_table(ws, 7 + len(file_rows), ["Metric", "Value"], validation_rows)

    base = 8 + len(file_rows) + len(validation_rows) + 1
    ws.cell(row=base, column=1, value="Cleaning Summary").font = SECTION_FONT
    cleaning_rows = [
        ["Total Rows Loaded", cleaning.rows_before],
        ["Duplicate Rows Removed", cleaning.duplicate_rows_removed],
        ["Invalid Rows Removed", cleaning.invalid_rows_removed],
        ["Missing Customer Names Filled", cleaning.missing_customer_names_filled],
        ["Missing Discounts Filled (set to 0)", cleaning.missing_discounts_filled],
        ["Whitespace Trimmed (cells)", cleaning.whitespace_trimmed_cells],
        ["Dates Normalized (rows)", cleaning.dates_normalized],
        ["Cancelled Orders Kept (excluded from revenue)", cleaning.cancelled_rows_kept],
        ["Rows Included in Final Report", cleaning.rows_final],
        ["Rows Excluded", cleaning.rows_excluded],
    ]
    _write_table(ws, base + 1, ["Metric", "Value"], cleaning_rows)

    warn_base = base + 1 + len(cleaning_rows) + 2
    ws.cell(row=warn_base, column=1, value="Validation Warnings (details)").font = SECTION_FONT
    warnings = [w.message for w in validation.warnings]
    if warnings:
        _write_table(ws, warn_base + 1, ["Warning"], [[w] for w in warnings])
    else:
        ws.cell(row=warn_base + 1, column=1, value="No warnings.").font = BODY_FONT

    _set_widths(ws, {"A": 44, "B": 24, "C": 18, "D": 12, "E": 40})


def _build_raw_data(ws, cleaned: pd.DataFrame, config: Config) -> None:
    _write_title(ws, "Raw Data", f"{config.company_name} | Cleaned and normalized sales transactions", end_col=18)
    currency = _currency_format(config)
    headers = list(cleaned.columns)
    first_data = _write_headers(ws, 4, headers)
    currency_cols = {"Unit Price", "Gross Revenue", "Discount Amount", "Net Revenue"}
    pct_cols = {"Discount"}
    for col_idx, col in enumerate(headers):
        for r in range(first_data, first_data + len(cleaned)):
            value = cleaned.iloc[r - first_data][col]
            cell = ws.cell(row=r, column=col_idx + 1, value=value)
            cell.font = BODY_FONT
            if col in currency_cols:
                cell.number_format = currency
            elif col in pct_cols:
                cell.number_format = "0.0%"
            elif col == "Order Date" and pd.notna(value):
                if isinstance(value, (pd.Timestamp, datetime)):
                    cell.number_format = "yyyy-mm-dd"
    _add_autofilter(ws, 4, headers, len(cleaned))
    _freeze_header(ws)
    widths = {"A": 16, "B": 12, "C": 12, "D": 24, "E": 12, "F": 30, "G": 18,
              "H": 10, "I": 12, "J": 10, "K": 18, "L": 16, "M": 16, "N": 12,
              "O": 16, "P": 14, "Q": 14, "R": 14}
    _set_widths(ws, widths)


# ---------------------------------------------------------------- entry

def generate_report(
    config: Config,
    cleaned_df: pd.DataFrame,
    file_infos: list[FileInfo],
    validation: ValidationResult,
    cleaning_report: CleaningReport,
    analytics_result: AnalyticsResult,
) -> Path:
    """Generate the professional workbook and return its path."""
    try:
        config.output_folder.mkdir(parents=True, exist_ok=True)
        today = datetime.now().strftime("%Y-%m-%d")
        filename = config.report_filename_pattern.format(date=today)
        report_path = config.output_folder / filename

        wb = Workbook()
        dashboard = wb.active
        dashboard.title = "Dashboard"

        _build_dashboard(dashboard, config, analytics_result)
        _build_executive_summary(
            wb.create_sheet("Executive Summary"), config, analytics_result, file_infos
        )
        _build_monthly(wb.create_sheet("Monthly Performance"), analytics_result, config)
        _build_product(wb.create_sheet("Product Performance"), analytics_result, config)
        _build_salesperson(wb.create_sheet("Salesperson Performance"), analytics_result, config)
        _build_region(wb.create_sheet("Regional Performance"), analytics_result, config)
        _build_customer(wb.create_sheet("Customer Performance"), analytics_result, config)
        _build_data_quality(
            wb.create_sheet("Data Quality"), config, file_infos, validation, cleaning_report
        )
        if config.include_raw_data:
            _build_raw_data(wb.create_sheet("Raw Data"), cleaned_df, config)

        wb.save(report_path)
        logger.info("Report saved to %s", report_path)
        return report_path
    except Exception as exc:
        from src.exceptions import ReportGenerationError

        logger.exception("Report generation failed")
        raise ReportGenerationError(f"Could not generate report: {exc}") from exc
